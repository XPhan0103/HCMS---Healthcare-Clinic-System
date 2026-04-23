import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

file_path = r'e:\Documents\Ky_8\Co Dieu AISDLC\HCMS---Healthcare-Clinic-System\Testing\01_UT\HCMS_Unit_Test_v0.1.xlsx'

def update_sheet(ws, method_name, test_cases):
    # Unmerge all cells first to avoid Read-Only errors
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    # Clear existing content
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    # Header Styles
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    bold_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Row 1: Header info
    ws['A1'] = "Code Module"
    ws['B1'] = "Patient"
    ws['F1'] = "Method"
    ws['I1'] = method_name
    
    ws['A2'] = "Created By"
    ws['B2'] = "SeniorQA"
    ws['F2'] = "Executed By"
    ws['I2'] = "SeniorQA"

    ws['A3'] = "Test requirement"
    ws['B3'] = f"Unit test for {method_name} logic in PatientService"

    ws['A4'] = "Passed"
    ws['C4'] = "Failed"
    ws['E4'] = "Untested"
    ws['G4'] = "N/A/B"
    ws['I4'] = "Total Test Cases"
    
    ws['A5'] = 0
    ws['C5'] = 0
    ws['E5'] = len(test_cases)
    ws['G5'] = 0
    ws['I5'] = len(test_cases)

    # Row 6: Test Case IDs
    ws['A6'] = "Condition"
    ws['B6'] = "Precondition"
    ws['C6'] = "Data Input"
    
    for i, tc in enumerate(test_cases):
        col = 6 + i # Column F, G, H...
        ws.cell(row=6, column=col).value = tc['id']
        ws.cell(row=6, column=col).font = bold_font

    # Mock Data Rows (Simplified Decision Table)
    ws['A7'] = "Input Data"
    for i, tc in enumerate(test_cases):
        col = 6 + i
        ws.cell(row=7, column=col).value = str(tc['input'])

    ws['A8'] = "Confirm"
    ws['B8'] = "Return"
    for i, tc in enumerate(test_cases):
        col = 6 + i
        ws.cell(row=8, column=col).value = tc['expected']

    ws['A9'] = "Result"
    for i, tc in enumerate(test_cases):
        col = 6 + i
        ws.cell(row=9, column=col).value = "PENDING"

    # Formatting
    for r in range(1, 10):
        for c in range(1, 6 + len(test_cases)):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

test_data = {
    'Create Patient': [
        {'id': 'UTCID01', 'input': 'Valid Data (Van A, 0901...)', 'expected': '201 Created + Response'},
        {'id': 'UTCID02', 'input': 'Short Name (An)', 'expected': '201 Created'},
        {'id': 'UTCID03', 'input': 'Empty Name', 'expected': '400 Validation Error'},
        {'id': 'UTCID04', 'input': 'Future DOB', 'expected': '400 Validation Error'},
        {'id': 'UTCID05', 'input': 'Invalid Phone (9 digits)', 'expected': '400 Validation Error'}
    ],
    'getPatientById': [
        {'id': 'UTCID08', 'input': 'Existing UUID', 'expected': '200 OK + Patient Data'},
        {'id': 'UTCID09', 'input': 'Non-existing UUID', 'expected': '404 EntityNotFound'},
        {'id': 'UTCID10', 'input': 'Null ID', 'expected': '400 IllegalArgument'}
    ],
    'updatePatientProfile': [
        {'id': 'UTCID11', 'input': 'Existing ID + New Height/Weight', 'expected': '200 OK + Updated Data'},
        {'id': 'UTCID12', 'input': 'Existing ID + Empty Allergies', 'expected': '200 OK (Clear Allergies)'},
        {'id': 'UTCID13', 'input': 'Non-existing ID', 'expected': '404 EntityNotFound'}
    ],
    'searchPatientsByPhone': [
        {'id': 'UTCID14', 'input': 'Phone: 0901234567 (Multi match)', 'expected': '200 OK + List[2]'},
        {'id': 'UTCID15', 'input': 'Phone: 0000000000 (No match)', 'expected': '200 OK + Empty List'},
        {'id': 'UTCID16', 'input': 'Empty Phone', 'expected': '200 OK + Empty List'}
    ]
}

wb = load_workbook(file_path)
for sheet_name, cases in test_data.items():
    if sheet_name in wb.sheetnames:
        update_sheet(wb[sheet_name], sheet_name, cases)

wb.save(file_path)
print("Updated all sheets successfully.")
